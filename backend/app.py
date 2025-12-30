from __future__ import annotations

import os
import time
from typing import Any, Dict, List, Optional, Tuple

from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from dateutil.parser import parse as dtparse
from openpyxl import load_workbook


# -----------------------------
# Config
# -----------------------------
EXCEL_PATH = os.getenv(
    "EXCEL_PATH",
    os.path.join(os.path.dirname(__file__), "data", "cases.xlsx")
)

SHEET_CASES = "CASES"
SHEET_SUSPECTS = "SUSPECTS"
SHEET_SEIZURES = "SEIZURES"
SHEET_LOOKUPS = "LOOKUPS"

# -----------------------------
# Column mappings (รองรับหลายชื่อ)
# -----------------------------
CASE_ID_KEYS: Tuple[str, ...] = ("Case ID (เลขประจำเคส)", "Case ID", "Case_ID")
CASE_NAME_KEYS: Tuple[str, ...] = ("ชื่อเคสคดี", "Case_Name")
CASE_TYPE_KEYS: Tuple[str, ...] = ("ประเภทคดี", "Case_Type")
CASE_DATE_KEYS: Tuple[str, ...] = ("วันที่รับคดี/เริ่มดำเนินการ", "Date_Received")
CASE_PLATFORM_KEYS: Tuple[str, ...] = ("แพลตฟอร์มหลัก", "Platform")
CASE_DAMAGE_KEYS: Tuple[str, ...] = ("มูลค่าความเสียหาย (บาท)", "Damage_Value_THB")
CASE_STATUS_KEYS: Tuple[str, ...] = ("สถานะคดี", "Case_Status")

SUS_CASE_ID_KEYS: Tuple[str, ...] = ("Case ID", "Case_ID")
SUS_NATIONALITY_KEYS: Tuple[str, ...] = ("สัญชาติ", "Nationality")
SUS_ARREST_KEYS: Tuple[str, ...] = ("สถานะการจับกุม", "Arrest_Status")
SUS_AGE_KEYS: Tuple[str, ...] = ("อายุ", "Age")

SEZ_CASE_ID_KEYS: Tuple[str, ...] = ("Case ID", "Case_ID")
SEZ_MAIN_KEYS: Tuple[str, ...] = ("ประเภทหลัก (3 ประเภท)", "Seizure_Category")
SEZ_ITEM_KEYS: Tuple[str, ...] = ("ประเภทสิ่งของ", "Item_Type", "ItemType")
SEZ_QTY_KEYS: Tuple[str, ...] = ("จำนวน", "Quantity")
SEZ_VALUE_KEYS: Tuple[str, ...] = ("มูลค่าโดยประมาณ (บาท)", "Value_THB")

# LOOKUPS: รองรับ 2 แบบ
# แบบเดิม (ธงประเทศ):
LOOKUP_FLAG_KEYS: Tuple[str, ...] = ("Flag",)
LOOKUP_COUNTRY_TH_KEYS: Tuple[str, ...] = ("ประเทศ (ภาษาไทย)", "Country_TH")

# แบบใหม่ (หน่วยของกลาง):
LOOKUP_ITEMTYPE_KEYS: Tuple[str, ...] = ("ItemType", "ประเภทสิ่งของ", "Item_Type")
LOOKUP_UNIT_KEYS: Tuple[str, ...] = ("Unit", "หน่วย")


# -----------------------------
# Helpers
# -----------------------------
def to_text(v: Any) -> str:
    return "" if v is None else str(v).strip()

def to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return 0.0

def to_int(v: Any) -> int:
    try:
        return int(round(to_float(v)))
    except Exception:
        return 0

def get_any(row: Dict[str, Any], keys: Tuple[str, ...], default: Any = None) -> Any:
    for k in keys:
        if k in row:
            return row.get(k)
    return default

def parse_date(v: Any) -> Optional[str]:
    if v is None or v == "":
        return None
    try:
        if hasattr(v, "strftime"):
            return v.strftime("%Y-%m-%d")
    except Exception:
        pass
    s = to_text(v)
    if not s:
        return None
    for dayfirst in (False, True):
        try:
            d = dtparse(s, dayfirst=dayfirst)
            return d.strftime("%Y-%m-%d")
        except Exception:
            continue
    return None

def in_date_range(date_str: Optional[str], date_from: Optional[str], date_to: Optional[str]) -> bool:
    if not date_from and not date_to:
        return True
    if not date_str:
        return False
    d = dtparse(date_str).date()
    if date_from and d < dtparse(date_from).date():
        return False
    if date_to and d > dtparse(date_to).date():
        return False
    return True

def contains(hay: str, needle: Optional[str]) -> bool:
    if not needle:
        return True
    return needle.lower() in (hay or "").lower()

def sheet_to_dicts(wb, sheet_name: str) -> List[Dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [to_text(h) for h in rows[0]]
    data: List[Dict[str, Any]] = []
    for r in rows[1:]:
        if r is None:
            continue
        obj: Dict[str, Any] = {}
        empty = True
        for i, h in enumerate(headers):
            if not h:
                continue
            val = r[i] if i < len(r) else None
            if val not in (None, ""):
                empty = False
            obj[h] = val
        if not empty:
            data.append(obj)
    return data

def _excel_mtime() -> Optional[float]:
    try:
        return os.path.getmtime(EXCEL_PATH)
    except Exception:
        return None


# -----------------------------
# Manual cache (สำคัญ: กัน reload พังแล้ว API ล่ม)
# -----------------------------
_CACHE: Dict[str, Any] = {
    "mtime": None,
    "data": None,
    "last_load_ts": None,
    "last_error": None,
}

def load_excel_uncached() -> Dict[str, Any]:
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Excel not found: {EXCEL_PATH}")

    wb = load_workbook(EXCEL_PATH, data_only=True)

    cases_raw = sheet_to_dicts(wb, SHEET_CASES)
    suspects_raw = sheet_to_dicts(wb, SHEET_SUSPECTS)
    seizures_raw = sheet_to_dicts(wb, SHEET_SEIZURES)
    lookups_raw = sheet_to_dicts(wb, SHEET_LOOKUPS)

    cases = []
    for r in cases_raw:
        cid = to_text(get_any(r, CASE_ID_KEYS))
        if not cid:
            continue
        cases.append({
            "case_id": cid,
            "case_name": to_text(get_any(r, CASE_NAME_KEYS)),
            "case_type": to_text(get_any(r, CASE_TYPE_KEYS)),
            "start_date": parse_date(get_any(r, CASE_DATE_KEYS)),
            "platform": to_text(get_any(r, CASE_PLATFORM_KEYS)),
            "damage": to_float(get_any(r, CASE_DAMAGE_KEYS)),
            "status": to_text(get_any(r, CASE_STATUS_KEYS)),
        })

    suspects = []
    for r in suspects_raw:
        cid = to_text(get_any(r, SUS_CASE_ID_KEYS))
        if not cid:
            continue
        suspects.append({
            "case_id": cid,
            "nationality": to_text(get_any(r, SUS_NATIONALITY_KEYS)) or "ไม่ระบุ",
            "arrest_status": to_text(get_any(r, SUS_ARREST_KEYS)),
            "age": to_int(get_any(r, SUS_AGE_KEYS)),
        })

    seizures = []
    for r in seizures_raw:
        cid = to_text(get_any(r, SEZ_CASE_ID_KEYS))
        if not cid:
            continue
        seizures.append({
            "case_id": cid,
            "main_category": to_text(get_any(r, SEZ_MAIN_KEYS)) or "ไม่ระบุ",
            "item_type": to_text(get_any(r, SEZ_ITEM_KEYS)) or "ไม่ระบุ",
            "qty": to_float(get_any(r, SEZ_QTY_KEYS)),
            "value": to_float(get_any(r, SEZ_VALUE_KEYS)),
        })

    # LOOKUPS - แบบธงประเทศ (ถ้ามี)
    flag_map: Dict[str, str] = {}
    for r in lookups_raw:
        country = to_text(get_any(r, LOOKUP_COUNTRY_TH_KEYS))
        flag = to_text(get_any(r, LOOKUP_FLAG_KEYS))
        if country and flag:
            flag_map[country] = flag

    # LOOKUPS - แบบหน่วยของกลาง (ถ้ามี)
    unit_map: Dict[str, str] = {}
    for r in lookups_raw:
        it = to_text(get_any(r, LOOKUP_ITEMTYPE_KEYS))
        u = to_text(get_any(r, LOOKUP_UNIT_KEYS))
        if it and u:
            unit_map[it] = u

    return {"cases": cases, "suspects": suspects, "seizures": seizures, "flag_map": flag_map, "unit_map": unit_map}

def ensure_fresh(force: bool = False) -> None:
    m = _excel_mtime()
    need = force or (_CACHE["data"] is None) or (m is not None and m != _CACHE["mtime"])
    if not need:
        return

    try:
        data = load_excel_uncached()  # โหลดให้ผ่านก่อน
        _CACHE["data"] = data
        _CACHE["mtime"] = m
        _CACHE["last_load_ts"] = time.time()
        _CACHE["last_error"] = None
    except Exception as e:
        # ถ้าโหลดไม่ผ่าน ให้ “ไม่ล้ม” และใช้ data เก่าต่อ
        _CACHE["last_error"] = str(e)


# -----------------------------
# FastAPI
# -----------------------------
app = FastAPI(title="TCSD2 Case Dashboard API", version="1.0.3")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get("/health")
def health():
    ensure_fresh()
    return {
        "status": "ok" if _CACHE["data"] is not None else "no-data",
        "excel": EXCEL_PATH,
        "mtime": _CACHE["mtime"],
        "last_error": _CACHE["last_error"],
    }

@app.get("/status")
def status():
    return dict(_CACHE)

@app.post("/reload")
def reload_excel():
    ensure_fresh(force=True)
    return {"status": "reloaded", **dict(_CACHE)}

@app.get("/dashboard")
def dashboard(
    q: Optional[str] = Query(default=None),
    case_id: Optional[str] = Query(default=None),
    date_from: Optional[str] = Query(default=None),
    date_to: Optional[str] = Query(default=None),
):
    ensure_fresh()

    if _CACHE["data"] is None:
        # ยังไม่มีข้อมูลเลย
        return {
            "kpis": {"total_cases": 0, "total_damage": 0, "total_suspects": 0, "arrested": 0, "not_arrested": 0},
            "arrest_donut": [],
            "nationalities": [],
            "seizures": [],
            "cases": [],
            "filters": {"q": q, "case_id": case_id, "date_from": date_from, "date_to": date_to},
            "meta": {"excel_mtime": _CACHE["mtime"], "last_error": _CACHE["last_error"]},
        }

    data = _CACHE["data"]
    cases = data["cases"]
    suspects = data["suspects"]
    seizures = data["seizures"]
    flag_map = data.get("flag_map", {})

    filtered_cases = []
    for c in cases:
        if not in_date_range(c.get("start_date"), date_from, date_to):
            continue
        if case_id and not contains(c.get("case_id", ""), case_id):
            continue
        if q and not (
            contains(c.get("case_id", ""), q)
            or contains(c.get("case_name", ""), q)
            or contains(c.get("case_type", ""), q)
            or contains(c.get("platform", ""), q)
        ):
            continue
        filtered_cases.append(c)

    case_ids = {c["case_id"] for c in filtered_cases}

    total_cases = len(filtered_cases)
    total_damage = sum(c.get("damage", 0.0) for c in filtered_cases)

    filtered_suspects = [s for s in suspects if s.get("case_id") in case_ids]
    total_suspects = len(filtered_suspects)
    arrested = sum(1 for s in filtered_suspects if s.get("arrest_status") == "จับกุมแล้ว")
    not_arrested = total_suspects - arrested

    nat_count: Dict[str, int] = {}
    for s in filtered_suspects:
        nat = to_text(s.get("nationality")) or "ไม่ระบุ"
        nat_count[nat] = nat_count.get(nat, 0) + 1

    nat_rows = []
    for country, cnt in sorted(nat_count.items(), key=lambda x: x[1], reverse=True):
        nat_rows.append({"flag": flag_map.get(country, "🏳️"), "country": country, "count": cnt})

    by_main_item: Dict[str, Dict[str, Dict[str, float]]] = {}
    filtered_seizures = [z for z in seizures if z.get("case_id") in case_ids]
    for z in filtered_seizures:
        main = to_text(z.get("main_category")) or "ไม่ระบุ"
        item = to_text(z.get("item_type")) or "ไม่ระบุ"
        by_main_item.setdefault(main, {})
        by_main_item[main].setdefault(item, {"qty": 0.0, "value": 0.0})
        by_main_item[main][item]["qty"] += float(z.get("qty", 0.0))
        by_main_item[main][item]["value"] += float(z.get("value", 0.0))

    def build_table(main_name: str) -> Dict[str, Any]:
        items = by_main_item.get(main_name, {})
        rows = [{"item_type": k, "qty": v["qty"], "value": v["value"]} for k, v in items.items()]
        rows.sort(key=lambda r: r["value"], reverse=True)
        return {
            "main_category": main_name,
            "rows": rows,
            "total_qty": sum(r["qty"] for r in rows),
            "total_value": sum(r["value"] for r in rows),
        }

    seizure_tables = [
        build_table("ตรวจยึดของกลาง"),
        build_table("ตรวจยึดสิ่งของเพื่อตรวจสอบความผิด"),
        build_table("ตรวจยึดทรัพย์สินที่เกี่ยวกับการฟอกเงิน"),
    ]

    cases_list = [{
        "case_id": c.get("case_id", ""),
        "case_name": c.get("case_name", ""),
        "case_type": c.get("case_type", ""),
        "start_date": c.get("start_date"),
        "platform": c.get("platform", ""),
        "damage": c.get("damage", 0.0),
        "status": c.get("status", ""),
    } for c in filtered_cases]

    return {
        "kpis": {
            "total_cases": total_cases,
            "total_damage": total_damage,
            "total_suspects": total_suspects,
            "arrested": arrested,
            "not_arrested": not_arrested,
        },
        "arrest_donut": [
            {"label": "จับกุมแล้ว", "value": arrested},
            {"label": "ยังไม่จับกุม", "value": not_arrested},
        ],
        "nationalities": nat_rows,
        "seizures": seizure_tables,
        "cases": cases_list,
        "filters": {"q": q, "case_id": case_id, "date_from": date_from, "date_to": date_to},
        "meta": {"excel_mtime": _CACHE["mtime"], "last_error": _CACHE["last_error"]},
    }
